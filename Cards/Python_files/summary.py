import xlrd
import xlsxwriter
from openpyxl import Workbook, load_workbook
from contextlib import closing
from win32com.client import Dispatch
import sys


def make_excel_file(file_name):
    with closing(Workbook()) as wb:
        wb.save(file_name)

def extract(filename,column_no): 
    loc = filename
    test=[]
    wb = xlrd.open_workbook(loc) 
    sheet = wb.sheet_by_index(0) 
    sheet.cell_value(0, 0) 
    
    for i in range(sheet.nrows): 
        test.append(sheet.cell_value(i, int(column_no)))
    return test

class ExcelUtility():
    def __init__(self,value):
        self.value = value
        print ("write to excel file"+str(self.value))

        
    def group(self,lst, n):
        return zip(*[lst[i::n] for i in range(n)])
    
    def write_to_excel_file(self,filename,sheetname,content_list):
            # Create an new Excel file and add a self.worksheet.
            self.workbook = xlsxwriter.Workbook(filename)
            self.worksheet=[]
            self.sheetname=sheetname
            for i in range (0,len(sheetname)):
                self.worksheet.append(self.workbook.add_worksheet(sheetname[i]))
                if sheetname[i]=='input':
                    self.worksheet[i].set_column('A:A',20)
                    self.worksheet[i].set_column('B:B',50)
                    self.worksheet[i].set_column('C:J',20)
                elif self.value==0:
                    self.worksheet[i].set_column('A:B', 50)
                    self.worksheet[i].set_column('C:D',20)
                    self.worksheet[i].set_column('E:E',40)
                    self.worksheet[i].set_column('F:H',50)
                    self.worksheet[i].set_column('I:I',20)
                else:
                    self.worksheet[i].set_column('A:A', 50)
                    self.worksheet[i].set_column('B:B', 20)
                    self.worksheet[i].set_column('C:C', 50)
                    self.worksheet[i].set_column('D:E',20)
                    self.worksheet[i].set_column('F:F',40)
                    self.worksheet[i].set_column('G:I',50)
                    self.worksheet[i].set_column('J:J',20)
                self.cell_format = self.workbook.add_format({'align':'top', 'border':1 , 'border_color':'black'})
                self.cell_format.set_text_wrap()
                self.bold = self.workbook.add_format({'bold': True , 'align':'top' , 'align':'center' , 'bg_color':'yellow' , 'font_size':14 , 'border':2 , 'border_color':'black'})
                self.bold.set_text_wrap()
                self.redbold=self.workbook.add_format({'bold': True, 'font_color': 'red', 'align':'top', 'border':1 , 'border_color':'black'})
                self.redbold.set_text_wrap()
                self.greenbold=self.workbook.add_format({'bold': True, 'font_color': 'green', 'align':'top', 'border':1 , 'border_color':'black'})
                self.greenbold.set_text_wrap()
                self.blue=self.workbook.add_format({'font_color': 'blue', 'align':'top', 'border':1 , 'border_color':'black'})
                self.blue.set_text_wrap()
                self.column0=self.workbook.add_format({'bold': True, 'align':'left' , 'align':'vcenter' , 'bg_color':'#b19cd9', 'border':1 , 'border_color':'black' , 'font_size': 12})
                self.column0.set_text_wrap()
                self.column1=self.workbook.add_format({'bold': True, 'align':'left' , 'align':'vcenter' , 'bg_color':'cyan', 'border':1 , 'border_color':'black' , 'font_size': 12})
                self.column1.set_text_wrap()
                self.Head=self.workbook.add_format({'bold': True, 'align':'left' , 'align':'vcenter' , 'bg_color':'#7CFC00', 'border':1 , 'border_color':'black' , 'font_size': 12})
                self.Head.set_text_wrap()
                #content_list=[1,1,'hello',2,1,'brother',3,1,'how are you',4,1,'are you good today']
                t=self.group(content_list[i],4)
                for item in t:
                    if item[3]==False:
                        self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.cell_format)
                    if item[3]=='Bold':
                        self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.bold)
                    elif item[3]=='Red_Bold':
                        self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.redbold)
                    elif item[3]=='Green_Bold':
                        self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.greenbold)
                    elif item[3]=='Blue':
                        self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.blue)
                    elif item[3]=='Column1':
                        self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.column1)
                    elif item[3]=='Column0':
                        self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.column0)
                    elif item[3]=='head':
                        self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.Head)
                # close work book
            self.workbook.close()
            print(self.worksheet)

file = open('config.txt')
data=[]
inc=-1
for each in file:
    inc+=1
    print (each)
    word = each.split()
    data.append([])
    for every in word:
        data[inc].append(every)
if data[inc][0]=='output_path=':
    reference=data[inc][1]
        
input_file=reference+'\\Audit_'+sys.argv[2]+'\\Audit_report_unique_'+sys.argv[2]+'.xlsx'
output_file1=reference+'\\Audit_'+sys.argv[2]+'\\Audit_report_summary_'+sys.argv[2]+'.xlsx'
output_file2=reference+'\\Audit_'+sys.argv[2]+'\\Audit_report_summary_alternate_'+sys.argv[2]+'.xlsx'
make_excel_file(output_file1)
make_excel_file(output_file2)
Part=extract(input_file,0)
Product=extract(input_file,1)
Date=extract(input_file,2)
Version=extract(input_file,3)
Os=extract(input_file,4)
File_name = extract(input_file,5)
Download_page=extract(input_file,6)
Description=extract(input_file,7)
Severity=extract(input_file,8)


search = extract('Python_files\\search.xlsx',0)
group = extract('Python_files\\search.xlsx',1)
short_group = extract('Python_files\\search.xlsx',2)

Part2 = extract(sys.argv[1],0)
card_name = extract(sys.argv[1],1)
Chip = extract(sys.argv[1],2)
Type = extract(sys.argv[1],3)
OFED_support = extract(sys.argv[1],4)
WinOF_support= extract(sys.argv[1],5)
WinOF2_support= extract(sys.argv[1],6)
VM_support = extract(sys.argv[1],7)
MFT_support = extract(sys.argv[1],8)
Windows_fw_support = extract(sys.argv[1],9)
Linux_RoCE_support = extract(sys.argv[1],10)
FW_binary_support = extract(sys.argv[1],11)


content1=[]
content2=[]
part_list=[]
group_name=[]
search_elem=[]
short_group_name=[]
for i in Part:
    if i=='Part Number':
        pass
    elif i not in part_list:
        part_list.append(i)
print(part_list)
for i in range (0, len(search)):
    if search[i]=='Keyword' or group[i]=='Group Name' or short_group[i]=='Group Short Name':
        pass
    elif search[i] not in search_elem:
        search_elem.append(search[i])
        group_name.append(group[i])
        short_group_name.append(short_group[i])
        
value=False

for every in range (0,len(part_list)):
    value='Bold'
    content1.append([])
    inc=0
    content1[every].append(inc)
    content1[every].append(0)
    content1[every].append(part_list[every])
    content1[every].append('head')
    content1[every].append(inc)
    content1[every].append(1)
    content1[every].append(card_name[every+1])
    content1[every].append('head')
    inc+=1
    content1[every].append(inc)
    content1[every].append(0)
    content1[every].append('Group Name')
    content1[every].append(value)
    content1[every].append(inc)
    content1[every].append(1)
    content1[every].append('Product Name')
    content1[every].append(value)
    content1[every].append(inc)
    content1[every].append(2)
    content1[every].append('Date')
    content1[every].append(value)
    content1[every].append(inc)
    content1[every].append(3)
    content1[every].append('Version')
    content1[every].append(value)
    content1[every].append(inc)
    content1[every].append(4)
    content1[every].append('OS')
    content1[every].append(value)
    content1[every].append(inc)
    content1[every].append(5)
    content1[every].append('File name')
    content1[every].append(value)
    content1[every].append(inc)
    content1[every].append(6)
    content1[every].append('Download_URL')
    content1[every].append(value)
    content1[every].append(inc)
    content1[every].append(7)
    content1[every].append('Description')
    content1[every].append(value)
    content1[every].append(inc)
    content1[every].append(8)
    content1[every].append('Severity')
    content1[every].append(value)
    for i in range (0,len(group_name)):
        flag=True
        value=False
        status=False
        for j in range (0,len(Product)):
                #print("Hello")
            if group_name[i]=='Mellanox OFED' and OFED_support[every+1]=='NO':
                status=True
                break
            elif group_name[i]=='WinOF' and WinOF_support[every+1]=='NO' :
                status=True
                break
            elif group_name[i]=='WinOF2' and WinOF2_support[every+1]=='NO':
                status=True
                break
            elif 'Mellanox MFT' in group_name[i] and MFT_support[every+1]=='NO':
                status=True
                break            
            elif 'ESXi' in group_name[i] and VM_support[every+1]=='NO':
                status=True
                break
            elif 'Windows firmware' in group_name[i] and Windows_fw_support[every+1]=='NO':
                status=True
                break
            elif 'Linux RoCE' in group_name[i] and Linux_RoCE_support[every+1]=='NO':
                status=True
                break
            elif 'Firmware binary' in group_name[i] and FW_binary_support[every+1]=='NO':
                status=True
                break
            elif search_elem[i] in Product[j] and part_list[every]==Part[j]:
                inc+=1 
                content1[every].append(inc)
                content1[every].append(0)
                if flag==True:
                    content1[every].append(group_name[i])
                    flag=False
                else:
                    content1[every].append(' ')
                content1[every].append('Column1')
                content1[every].append(inc)
                content1[every].append(1)
                if group_name[i]=='Mellanox OFED':
                    P = Product[j].split('for')
                    content1[every].append(P[1])
                else:
                    content1[every].append(Product[j])
                if part_list[every] not in Product[j] :
                    if group_name[i]=='Firmware binary posting':
                        print(str(every))
                        print(part_list[every])
                        print(Product[j])
                        content1[every].append('Blue')
                    else:
                        content1[every].append(value)
                else:
                    content1[every].append(value)
                        
                content1[every].append(inc)
                content1[every].append(2)
                content1[every].append(Date[j])
                content1[every].append(value)
                content1[every].append(inc)
                content1[every].append(3)
                content1[every].append(Version[j])
                content1[every].append(value)
                content1[every].append(inc)
                content1[every].append(4)
                content1[every].append(Os[j])
                content1[every].append(value)
                content1[every].append(inc)
                content1[every].append(5)
                content1[every].append(File_name[j])
                if File_name[j]=='Not Found':
                    content1[every].append('Red_Bold')
                else:
                    content1[every].append(value)
                content1[every].append(inc)
                content1[every].append(6)
                content1[every].append(Download_page[j])
                content1[every].append(value)
                content1[every].append(inc)
                content1[every].append(7)
                content1[every].append(Description[j])
                content1[every].append(value)
                content1[every].append(inc)
                content1[every].append(8)
                content1[every].append(Severity[j])
                content1[every].append(value)
      
        if flag==True:
            inc+=1
            value = False
            content1[every].append(inc)
            content1[every].append(0)
            content1[every].append(group_name[i])
            content1[every].append('Column1')
            content1[every].append(inc)
            content1[every].append(1)
            if status==True:
                content1[every].append('Not Supported')
                content1[every].append('Green_Bold')
                status=False
            elif group_name[i]=='Firmware binary posting':
                content1[every].append('No firmware posting for '+part_list[every]+' found')
                content1[every].append('Red_Bold')
            else:
                content1[every].append('No Products Found')
                content1[every].append('Red_Bold')
            content1[every].append(inc)
            content1[every].append(2)
            content1[every].append(' ')
            content1[every].append(value)
            content1[every].append(inc)
            content1[every].append(3)
            content1[every].append(' ')
            content1[every].append(value)
            content1[every].append(inc)
            content1[every].append(4)
            content1[every].append(' ')
            content1[every].append(value)
            content1[every].append(inc)
            content1[every].append(5)
            content1[every].append(' ')
            content1[every].append(value)
            content1[every].append(inc)
            content1[every].append(6)
            content1[every].append(' ')
            content1[every].append(value)
            content1[every].append(inc)
            content1[every].append(7)
            content1[every].append(' ')
            content1[every].append(value)
            content1[every].append(inc)
            content1[every].append(8)
            content1[every].append(' ')
            content1[every].append(value)


value=False

for every in range (0,len(group_name)):
    value='Bold'
    content2.append([])
    inc=0
    content2[every].append(inc)
    content2[every].append(0)
    content2[every].append(group_name[every])
    content2[every].append('head')
    inc+=1
    content2[every].append(inc)
    content2[every].append(0)
    content2[every].append('Card Name')
    content2[every].append(value)
    content2[every].append(inc)
    content2[every].append(1)
    content2[every].append('Part Number')
    content2[every].append(value)
    content2[every].append(inc)
    content2[every].append(2)
    content2[every].append('Product Name')
    content2[every].append(value)
    content2[every].append(inc)
    content2[every].append(3)
    content2[every].append('Date')
    content2[every].append(value)
    content2[every].append(inc)
    content2[every].append(4)
    content2[every].append('Version')
    content2[every].append(value)
    content2[every].append(inc)
    content2[every].append(5)
    content2[every].append('OS')
    content2[every].append(value)
    content2[every].append(inc)
    content2[every].append(6)
    content2[every].append('File name')
    content2[every].append(value)
    content2[every].append(inc)
    content2[every].append(7)
    content2[every].append('Download_URL')
    content2[every].append(value)
    content2[every].append(inc)
    content2[every].append(8)
    content2[every].append('Description')
    content2[every].append(value)
    content2[every].append(inc)
    content2[every].append(9)
    content2[every].append('Severity')
    content2[every].append(value)
    for i in range (0,len(part_list)):
        flag=True
        value=False
        status=False
        for j in range (0,len(Product)):
                #print("Hello")
            if group_name[every]=='Mellanox OFED' and OFED_support[i+1]=='NO':
                
                status=True
                break
            elif group_name[every]=='WinOF' and WinOF_support[i+1]=='NO' :
                status=True
                
                break
            elif group_name[every]=='WinOF2' and WinOF2_support[i+1]=='NO':
                status=True
               
                break
            elif 'Mellanox MFT' in group_name[every] and MFT_support[i+1]=='NO':
                status=True
                break            
            elif 'ESXi' in group_name[every] and VM_support[i+1]=='NO':
                status=True
                break
            elif 'Windows firmware' in group_name[every] and Windows_fw_support[i+1]=='NO':
                status=True
                break
            elif 'Linux RoCE' in group_name[every] and Linux_RoCE_support[i+1]=='NO':
                status=True
                break
            elif 'Firmware binary' in group_name[every] and FW_binary_support[i+1]=='NO':
                status=True
                break
            elif search_elem[every] in Product[j] and part_list[i]==Part[j]:
                inc+=1 
                content2[every].append(inc)
                content2[every].append(0)
                if flag==True:
                    content2[every].append(card_name[i+1])
                    content2[every].append('Column0')
                    content2[every].append(inc)
                    content2[every].append(1)
                    content2[every].append(part_list[i])
                    flag=False
                else:
                    content2[every].append(' ')
                    content2[every].append('Column0')
                    content2[every].append(inc)
                    content2[every].append(1)
                    content2[every].append(' ')
                content2[every].append('Column1')
                content2[every].append(inc)
                content2[every].append(2)
                if group_name[every]=='Mellanox OFED':
                    P = Product[j].split('for')
                    content2[every].append(P[1])
                else:
                    content2[every].append(Product[j])
                if part_list[i] not in Product[j] and group_name[every]=='Firmware binary posting':
                    content2[every].append('Blue')
                else:
                    content2[every].append(value)
                content2[every].append(inc)
                content2[every].append(3)
                content2[every].append(Date[j])
                content2[every].append(value)
                content2[every].append(inc)
                content2[every].append(4)
                content2[every].append(Version[j])
                content2[every].append(value)
                content2[every].append(inc)
                content2[every].append(5)
                content2[every].append(Os[j])
                content2[every].append(value)
                content2[every].append(inc)
                content2[every].append(6)
                content2[every].append(File_name[j])
                if File_name[j]=='Not Found':
                    content2[every].append('Red_Bold')
                else:
                    content2[every].append(value) 
                content2[every].append(inc)
                content2[every].append(7)
                content2[every].append(Download_page[j])
                content2[every].append(value)
                content2[every].append(inc)
                content2[every].append(8)
                content2[every].append(Description[j])
                content2[every].append(value)
                content2[every].append(inc)
                content2[every].append(9)
                content2[every].append(Severity[j])
                content2[every].append(value)
     
        if flag==True:
            inc+=1
            value = False
            content2[every].append(inc)
            content2[every].append(0)
            content2[every].append(card_name[i+1])
            content2[every].append('Column0')
            content2[every].append(inc)
            content2[every].append(1)
            content2[every].append(part_list[i])
            content2[every].append('Column1')
            content2[every].append(inc)
            content2[every].append(2)
            if status==True:
                content2[every].append('Not Supported')
                content2[every].append('Green_Bold')
                status=False
            elif group_name[every]=='Firmware binary posting':
                content2[every].append('No firmware posting for '+part_list[i]+' found')
                content2[every].append('Red_Bold')
            else:
                content2[every].append('No Products Found')
                content2[every].append('Red_Bold')
  
            content2[every].append(inc)
            content2[every].append(3)
            content2[every].append(' ')
            content2[every].append(value)
            content2[every].append(inc)
            content2[every].append(4)
            content2[every].append(' ')
            content2[every].append(value)
            content2[every].append(inc)
            content2[every].append(5)
            content2[every].append(' ')
            content2[every].append(value)
            content2[every].append(inc)
            content2[every].append(6)
            content2[every].append(' ')
            content2[every].append(value)
            content2[every].append(inc)
            content2[every].append(7)
            content2[every].append(' ')
            content2[every].append(value)
            content2[every].append(inc)
            content2[every].append(8)
            content2[every].append(' ')
            content2[every].append(value)
            content2[every].append(inc)
            content2[every].append(9)
            content2[every].append(' ')
            content2[every].append(value)
add=[]
for i in range(0,len(Part2)):
    add.append(i)
    add.append(0)
    add.append(Part2[i])
    if i==0:
        add.append('Bold')
    else:
        add.append(False)
    add.append(i)
    add.append(1)
    add.append(card_name[i])
    if i==0:
        add.append('Bold')
    else:
        add.append(False)
    add.append(i)
    add.append(2)
    add.append(Chip[i])
    if i==0:
        add.append('Bold')
    else:
        add.append(False)
    add.append(i)
    add.append(3)
    add.append(Type[i])
    if i==0:
        add.append('Bold')
    else:
        add.append(False)
    add.append(i)
    add.append(4)
    add.append(WinOF_support[i])
    if i==0:
        add.append('Bold')
    else:
        add.append(False)
    add.append(i)
    add.append(5)
    add.append(WinOF2_support[i])
    if i==0:
        add.append('Bold')
    else:
        add.append(False)
    add.append(i)
    add.append(6)
    add.append(OFED_support[i])
    if i==0:
        add.append('Bold')
    else:
        add.append(False)
    add.append(i)
    add.append(7)
    add.append(VM_support[i])
    if i==0:
        add.append('Bold')
    else:
        add.append(False)
    add.append(i)
    add.append(8)
    add.append(MFT_support[i])
    if i==0:
        add.append('Bold')
    else:
        add.append(False)
    add.append(i)
    add.append(9)
    add.append(Windows_fw_support[i])
    if i==0:
        add.append('Bold')
    else:
        add.append(False)
    add.append(i)
    add.append(10)
    add.append(Linux_RoCE_support[i])
    if i==0:
        add.append('Bold')
    else:
        add.append(False)
    add.append(i)
    add.append(11)
    add.append(FW_binary_support[i])
    if i==0:
        add.append('Bold')
    else:
        add.append(False)
    
content1.insert(0,add)
content2.insert(0,add)

obj1=ExcelUtility(0)
obj2=ExcelUtility(1)
sheet_list1=[]
sheet_list2=[]
sheet_list1.append('input')
for i in range (0,len(part_list)):
    #print(content1[i])
    sheet_list1.append(part_list[i])
sheet_list2.append('input')
for i in range (0,len(short_group_name)):
    sheet_list2.append(short_group_name[i])

obj1.write_to_excel_file(output_file1,sheet_list1,content1)
obj2.write_to_excel_file(output_file2,sheet_list2,content2)