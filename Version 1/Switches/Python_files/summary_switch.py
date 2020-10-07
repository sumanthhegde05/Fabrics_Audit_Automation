import xlrd
import xlsxwriter
from openpyxl import Workbook, load_workbook
from contextlib import closing
from win32com.client import Dispatch
import sys


def make_excel_file(file_name):
    with closing(Workbook()) as wb:
        wb.save(file_name)

def extract(filename,column_no): 
    loc = filename
    test=[]
    wb = xlrd.open_workbook(loc) 
    sheet = wb.sheet_by_index(0) 
    sheet.cell_value(0, 0) 
    
    for i in range(sheet.nrows): 
        test.append(sheet.cell_value(i, int(column_no)))
    return test

class ExcelUtility():
    def __init__(self,value):
        self.value = value
        print ("write to excel file"+str(self.value))

        
    def group(self,lst, n):
        return zip(*[lst[i::n] for i in range(n)])
    
    def write_to_excel_file(self,filename,sheetname,content_list):
            # Create an new Excel file and add a self.worksheet.
            self.workbook = xlsxwriter.Workbook(filename)
            self.worksheet=[]
            self.sheetname=sheetname
            for i in range (0,len(sheetname)):
                self.worksheet.append(self.workbook.add_worksheet(self.sheetname[i]))
                if self.sheetname[i]=='input':
                    self.worksheet[i].set_column('A:A', 20)
                    self.worksheet[i].set_column('B:B',50)
                    self.worksheet[i].set_column('C:D',20)
                else:
                    self.worksheet[i].set_column('A:A',50)
                    self.worksheet[i].set_column('B:C',20)
                    self.worksheet[i].set_column('D:D',40)
                    self.worksheet[i].set_column('E:G',50)
                    self.worksheet[i].set_column('H:H',20)
                self.cell_format = self.workbook.add_format({'align':'top', 'border':1 , 'border_color':'black'})
                self.cell_format.set_text_wrap()
                self.bold = self.workbook.add_format({'bold': True , 'align':'top' , 'align':'center' , 'bg_color':'yellow' , 'font_size':14 , 'border':2 , 'border_color':'black'})
                self.bold.set_text_wrap()
                self.redbold=self.workbook.add_format({'bold': True, 'font_color': 'red', 'align':'top', 'border':1 , 'border_color':'black'})
                self.redbold.set_text_wrap()
                self.blue=self.workbook.add_format({'font_color': 'blue', 'align':'top', 'border':1 , 'border_color':'black'})
                self.blue.set_text_wrap()
                self.column0=self.workbook.add_format({'bold': True, 'align':'left' , 'align':'vcenter' , 'bg_color':'#b19cd9', 'border':1 , 'border_color':'black' , 'font_size': 12})
                self.column0.set_text_wrap()
                self.column1=self.workbook.add_format({'bold': True, 'align':'left' , 'align':'vcenter' , 'bg_color':'cyan', 'border':1 , 'border_color':'black' , 'font_size': 12})
                self.column1.set_text_wrap()
                self.Head=self.workbook.add_format({'bold': True, 'align':'left' , 'align':'vcenter' , 'bg_color':'#7CFC00', 'border':1 , 'border_color':'black' , 'font_size': 12})
                self.Head.set_text_wrap()
                #content_list=[1,1,'hello',2,1,'brother',3,1,'how are you',4,1,'are you good today']
                t=self.group(content_list[i],4)
                for item in t:
                    if item[3]==False:
                        self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.cell_format)
                    if item[3]=='Bold':
                        self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.bold)
                    elif item[3]=='Red_Bold':
                        self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.redbold)
                    elif item[3]=='Blue':
                        self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.blue)
                    elif item[3]=='Column1':
                        self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.column1)
                    elif item[3]=='Column0':
                        self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.column0)
                    elif item[3]=='head':
                        self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.Head)
                # close work book
            self.workbook.close()
            print(self.worksheet)

file = open('config.txt')
data=[]
inc=-1
for each in file:
    inc+=1
    print (each)
    word = each.split()
    data.append([])
    for every in word:
        data[inc].append(every)
if data[inc][0]=='output_path=':
    reference=data[inc][1]
        
input_file=reference+'\\Audit_'+sys.argv[2]+'\\Audit_report_unique_'+sys.argv[2]+'.xlsx'
output_file1=reference+'\\Audit_'+sys.argv[2]+'\\Audit_report_summary_'+sys.argv[2]+'.xlsx'

make_excel_file(output_file1)

Part=extract(input_file,0)
Card=extract(input_file,1)
Product=extract(input_file,2)
Date=extract(input_file,3)
Version=extract(input_file,4)
Os=extract(input_file,5)
File_name = extract(input_file,6)
Download_page=extract(input_file,7)
Description=extract(input_file,8)
Severity=extract(input_file,9)



Part2 = extract(sys.argv[1],0)
card_name = extract(sys.argv[1],1)
Chip = extract(sys.argv[1],2)
Type = extract(sys.argv[1],3)

content1=[]

part_list=[]
card_list=[]
for i in Part:
    if i=='Part Number':
        pass
    elif i not in part_list:
        part_list.append(i)
for j in Card:
    if j=='Marketing Name':
        pass
    elif j not in card_list:
        card_list.append(j)
        
value=False

for every in range (0,len(part_list)):
    value='Bold'
    content1.append([])
    inc=0
 
    content1[every].append(inc)
    content1[every].append(0)
    content1[every].append(card_list[every])
    content1[every].append('head')
    inc=1
    content1[every].append(inc)
    content1[every].append(0)
    content1[every].append('Product Name')
    content1[every].append(value)
    content1[every].append(inc)
    content1[every].append(1)
    content1[every].append('Date')
    content1[every].append(value)
    content1[every].append(inc)
    content1[every].append(2)
    content1[every].append('Version')
    content1[every].append(value)
    content1[every].append(inc)
    content1[every].append(3)
    content1[every].append('Os')
    content1[every].append(value)
    content1[every].append(inc)
    content1[every].append(4)
    content1[every].append('File name')
    content1[every].append(value)
    content1[every].append(inc)
    content1[every].append(5)
    content1[every].append('Download_page')
    content1[every].append(value)
    content1[every].append(inc)
    content1[every].append(6)
    content1[every].append('Description')
    content1[every].append(value)
    content1[every].append(inc)
    content1[every].append(7)
    content1[every].append('Severity')
    content1[every].append(value)
    for j in range (0,len(Product)):
        value=False
        if part_list[every]==Part[j]:
                inc+=1 
                content1[every].append(inc)
                content1[every].append(0)
                content1[every].append(Product[j])
                if part_list[every] not in Product[j]:
                    content1[every].append('Blue')
                else:
                    content1[every].append(value)
                content1[every].append(inc)
                content1[every].append(1)
                content1[every].append(Date[j])
                content1[every].append(value)
                content1[every].append(inc)
                content1[every].append(2)
                content1[every].append(Version[j])
                content1[every].append(value)
                content1[every].append(inc)
                content1[every].append(3)
                content1[every].append(Os[j])
                content1[every].append(value)
                content1[every].append(inc)
                content1[every].append(4)
                content1[every].append(File_name[j])
                if File_name[j]=='Not Found':
                    content1[every].append('Red_Bold')
                else:
                    content1[every].append(value)
                content1[every].append(inc)
                content1[every].append(5)
                content1[every].append(Download_page[j])
                content1[every].append(value)
                content1[every].append(inc)
                content1[every].append(6)
                content1[every].append(Description[j])
                content1[every].append(value)
                content1[every].append(inc)
                content1[every].append(7)
                content1[every].append(Severity[j])
                content1[every].append(value)
      
add=[]
for i in range(0,len(Part2)):
    add.append(i)
    add.append(0)
    add.append(Part2[i])
    if i==0:
        add.append('Bold')
    else:
        add.append(False)
    add.append(i)
    add.append(1)
    add.append(card_name[i])
    if i==0:
        add.append('Bold')
    else:
        add.append(False)
    add.append(i)
    add.append(2)
    add.append(Chip[i])
    if i==0:
        add.append('Bold')
    else:
        add.append(False)
    add.append(i)
    add.append(3)
    add.append(Type[i])
    if i==0:
        add.append('Bold')
    else:
        add.append(False)
    
content1.insert(0,add)


obj1=ExcelUtility(0)

sheet_list1=[]
sheet_list1.append('input')
for i in range (0,len(part_list)):
    #print(content1[i])
    sheet_list1.append(part_list[i])


obj1.write_to_excel_file(output_file1,sheet_list1,content1)
