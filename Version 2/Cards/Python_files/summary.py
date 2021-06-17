import xlrd
import xlsxwriter
from openpyxl import Workbook, load_workbook
from contextlib import closing
from win32com.client import Dispatch
import sys
import logging
import pandas as pd

def make_excel_file(file_name):
    with closing(Workbook()) as wb:
        wb.save(file_name)


def extract(filename,column_no): 
    print(filename)
    df = pd.ExcelFile(filename).parse(0)
    ret_list = df.iloc[:,int(column_no)].values.tolist()
    ret_list.insert(0,df.columns.values.tolist()[int(column_no)])
    return ret_list


class ExcelUtility():
    def __init__(self,value):
        self.value = value
    
    def group(self,lst, n):
        return zip(*[lst[i::n] for i in range(n)])
    
    
    def write_to_excel_file(self,filename,sheetname,content_list):
            self.workbook = xlsxwriter.Workbook(filename)
            self.worksheet=[]
            self.sheetname=sheetname
            for i in range (0,len(sheetname)):
                self.worksheet.append(self.workbook.add_worksheet(sheetname[i]))
                if sheetname[i]=='input':
                    self.worksheet[i].set_column('A:A',20)
                    self.worksheet[i].set_column('B:B',50)
                    self.worksheet[i].set_column('C:J',20)
                elif self.value==0:
                    self.worksheet[i].set_column('A:B', 50)
                    self.worksheet[i].set_column('C:D',20)
                    self.worksheet[i].set_column('E:E',40)
                    self.worksheet[i].set_column('F:H',50)
                    self.worksheet[i].set_column('I:I',20)
                else:
                    self.worksheet[i].set_column('A:A', 50)
                    self.worksheet[i].set_column('B:B', 20)
                    self.worksheet[i].set_column('C:C', 50)
                    self.worksheet[i].set_column('D:E',20)
                    self.worksheet[i].set_column('F:F',40)
                    self.worksheet[i].set_column('G:I',50)
                    self.worksheet[i].set_column('J:J',20)
                self.cell_format = self.workbook.add_format({'align':'top', 'border':1 , 'border_color':'black'})
                self.cell_format.set_text_wrap()
                self.bold = self.workbook.add_format({'bold': True , 'align':'top' , 'align':'center' , 'bg_color':'yellow' , 'font_size':14 , 'border':2 , 'border_color':'black'})
                self.bold.set_text_wrap()
                self.redbold=self.workbook.add_format({'bold': True, 'font_color': 'red', 'align':'top', 'border':1 , 'border_color':'black'})
                self.redbold.set_text_wrap()
                self.greenbold=self.workbook.add_format({'bold': True, 'font_color': 'green', 'align':'top', 'border':1 , 'border_color':'black'})
                self.greenbold.set_text_wrap()
                self.blue=self.workbook.add_format({'font_color': 'blue', 'align':'top', 'border':1 , 'border_color':'black'})
                self.blue.set_text_wrap()
                self.column0=self.workbook.add_format({'bold': True, 'align':'left' , 'align':'vcenter' , 'bg_color':'#b19cd9', 'border':1 , 'border_color':'black' , 'font_size': 12})
                self.column0.set_text_wrap()
                self.column1=self.workbook.add_format({'bold': True, 'align':'left' , 'align':'vcenter' , 'bg_color':'cyan', 'border':1 , 'border_color':'black' , 'font_size': 12})
                self.column1.set_text_wrap()
                self.Head=self.workbook.add_format({'bold': True, 'align':'left' , 'align':'vcenter' , 'bg_color':'#7CFC00', 'border':1 , 'border_color':'black' , 'font_size': 12})
                self.Head.set_text_wrap()
                temp = self.group(content_list[i],4)
                
                for item in temp:
                    try:
                        if item[3]==False:
                            self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.cell_format)
                        if item[3]=='Bold':
                            self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.bold)
                        elif item[3]=='Red_Bold':
                            self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.redbold)
                        elif item[3]=='Green_Bold':
                            self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.greenbold)
                        elif item[3]=='Blue':
                            self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.blue)
                        elif item[3]=='Column1':
                            self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.column1)
                        elif item[3]=='Column0':
                            self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.column0)
                        elif item[3]=='head':
                            self.worksheet[i].write(int(item[0]), int(item[1]), item[2],self.Head)
                    except:
                        print("error",item)
                        
                
            self.workbook.close()

def get_output_path():
    file = open('config.txt')
    data=[]
    inc=-1

    for each in file:
        inc+=1
        word = each.split()
        data.append([])
        for every in word:
            data[inc].append(every)
            
    if data[inc][0]=='output_path=':
        return data[inc][1]
    

def create_output_files(output_path):
    #print(output_path)
    input_file      =   output_path+'\\Audit_'+sys.argv[2]+'\\Audit_report_unique_'+sys.argv[2]+'.xlsx'
    output_file1    =   output_path+'\\Audit_'+sys.argv[2]+'\\Audit_report_summary_'+sys.argv[2]+'.xlsx'
    output_file2    =   output_path+'\\Audit_'+sys.argv[2]+'\\Audit_report_summary_alternate_'+sys.argv[2]+'.xlsx'
    make_excel_file(output_file1)
    make_excel_file(output_file2)
    return input_file, output_file1, output_file2


def get_keywords():
    search = extract('search.xlsx',0)
    group = extract('search.xlsx',1)
    short_group = extract('search.xlsx',2)
    return search, group, short_group


def get_logger():
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)
    file_handler = logging.FileHandler("debug_main.log",mode='a')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(logging.Formatter('%(name)s : %(levelname)-8s : %(lineno)s : %(message)s'))
    """console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter('%(message)s'))"""
    logger.addHandler(file_handler)
    #logger.addHandler(console_handler)
    return logger





def add_to_summary_content():
    global summary_content
    


def summary():
    global summary_content
    value=False
    
    for every in range (0,len(part_list)):
        value='Bold'
        summary_content.append([])
        inc=0
        summary_content[every].append(inc)
        summary_content[every].append(0)
        summary_content[every].append(part_list[every])
        summary_content[every].append('head')
        summary_content[every].append(inc)
        summary_content[every].append(1)
        
        summary_content[every].append(card_name[every+1])
        summary_content[every].append('head')
        inc+=1
        summary_content[every].append(inc)
        summary_content[every].append(0)
        summary_content[every].append('Group Name')
        summary_content[every].append(value)
        summary_content[every].append(inc)
        summary_content[every].append(1)
        summary_content[every].append('Product Name')
        summary_content[every].append(value)
        summary_content[every].append(inc)
        summary_content[every].append(2)
        summary_content[every].append('Date')
        summary_content[every].append(value)
        summary_content[every].append(inc)
        summary_content[every].append(3)
        summary_content[every].append('Version')
        summary_content[every].append(value)
        summary_content[every].append(inc)
        summary_content[every].append(4)
        summary_content[every].append('OS')
        summary_content[every].append(value)
        summary_content[every].append(inc)
        summary_content[every].append(5)
        summary_content[every].append('File name')
        summary_content[every].append(value)
        summary_content[every].append(inc)
        summary_content[every].append(6)
        summary_content[every].append('Download_URL')
        summary_content[every].append(value)
        summary_content[every].append(inc)
        summary_content[every].append(7)
        summary_content[every].append('Description')
        summary_content[every].append(value)
        summary_content[every].append(inc)
        summary_content[every].append(8)
        summary_content[every].append('Severity')
        summary_content[every].append(value)
        
        for i in range (0,len(group_name)):
            flag=True
            value=False
            status=False
            
            for j in range (0,len(Product_name)):
                
                if group_name[i]=='Mellanox OFED' and OFED_support[every+1]=='NO':
                    status=True
                    break
                elif group_name[i]=='WinOF' and WinOF_support[every+1]=='NO' :
                    status=True
                    break
                elif group_name[i]=='WinOF2' and WinOF2_support[every+1]=='NO':
                    status=True
                    break
                elif 'Mellanox MFT' in group_name[i] and MFT_support[every+1]=='NO':
                    status=True
                    break            
                elif 'ESXi' in group_name[i] and VM_support[every+1]=='NO':
                    status=True
                    break
                elif 'Windows firmware' in group_name[i] and Windows_fw_support[every+1]=='NO':
                    status=True
                    break
                elif 'Linux RoCE' in group_name[i] and Linux_RoCE_support[every+1]=='NO':
                    status=True
                    break
                elif 'Firmware binary' in group_name[i] and FW_binary_support[every+1]=='NO':
                    status=True
                    break
                elif search_elem[i] in Product_name[j] and part_list[every]==Part_numbers[j]:
                    inc+=1 
                    summary_content[every].append(inc)
                    summary_content[every].append(0)
                    
                    if flag==True:
                        summary_content[every].append(group_name[i])
                        flag=False
                    else:
                        summary_content[every].append(' ')
                        
                    summary_content[every].append('Column1')
                    summary_content[every].append(inc)
                    summary_content[every].append(1)
                    
                    if group_name[i]=='Mellanox OFED':
                        P = Product_name[j].split('for')
                        summary_content[every].append(P[1])
                    else:
                        summary_content[every].append(Product_name[j])
                        
                    if part_list[every].strip() not in Product_name[j] :
                        
                        if group_name[i]=='Firmware binary posting':
                            summary_content[every].append('Blue')
                        else:
                            summary_content[every].append(value)
                            
                    else:
                        summary_content[every].append(value)
                            
                    summary_content[every].append(inc)
                    summary_content[every].append(2)
                    summary_content[every].append(Date[j])
                    summary_content[every].append(value)
                    summary_content[every].append(inc)
                    summary_content[every].append(3)
                    summary_content[every].append(Version[j])
                    summary_content[every].append(value)
                    summary_content[every].append(inc)
                    summary_content[every].append(4)
                    summary_content[every].append(Os[j])
                    summary_content[every].append(value)
                    summary_content[every].append(inc)
                    summary_content[every].append(5)
                    summary_content[every].append(File_name[j])
                    
                    if File_name[j]=='Not Found':
                        summary_content[every].append('Red_Bold')
                    else:
                        summary_content[every].append(value)
                        
                    summary_content[every].append(inc)
                    summary_content[every].append(6)
                    summary_content[every].append(Download_page[j])
                    summary_content[every].append(value)
                    summary_content[every].append(inc)
                    summary_content[every].append(7)
                    summary_content[every].append(Description[j])
                    summary_content[every].append(value)
                    summary_content[every].append(inc)
                    summary_content[every].append(8)
                    summary_content[every].append(Severity[j])
                    summary_content[every].append(value)
        
            if flag==True:
                inc+=1
                value = False
                summary_content[every].append(inc)
                summary_content[every].append(0)
                summary_content[every].append(group_name[i])
                summary_content[every].append('Column1')
                summary_content[every].append(inc)
                summary_content[every].append(1)
                
                if status==True:
                    summary_content[every].append('Not Supported')
                    summary_content[every].append('Green_Bold')
                    status=False
                elif group_name[i]=='Firmware binary posting':
                    summary_content[every].append('No firmware posting for '+part_list[every]+' found')
                    summary_content[every].append('Red_Bold')
                else:
                    summary_content[every].append('No Products Found')
                    summary_content[every].append('Red_Bold')
                    
                summary_content[every].append(inc)
                summary_content[every].append(2)
                summary_content[every].append(' ')
                summary_content[every].append(value)
                summary_content[every].append(inc)
                summary_content[every].append(3)
                summary_content[every].append(' ')
                summary_content[every].append(value)
                summary_content[every].append(inc)
                summary_content[every].append(4)
                summary_content[every].append(' ')
                summary_content[every].append(value)
                summary_content[every].append(inc)
                summary_content[every].append(5)
                summary_content[every].append(' ')
                summary_content[every].append(value)
                summary_content[every].append(inc)
                summary_content[every].append(6)
                summary_content[every].append(' ')
                summary_content[every].append(value)
                summary_content[every].append(inc)
                summary_content[every].append(7)
                summary_content[every].append(' ')
                summary_content[every].append(value)
                summary_content[every].append(inc)
                summary_content[every].append(8)
                summary_content[every].append(' ')
                summary_content[every].append(value)


def summary_alt():
    global alt_summary_content
    value=False
    
    for every in range (0,len(group_name)):
        value='Bold'
        alt_summary_content.append([])
        inc=0
        alt_summary_content[every].append(inc)
        alt_summary_content[every].append(0)
        alt_summary_content[every].append(group_name[every])
        alt_summary_content[every].append('head')
        inc+=1
        alt_summary_content[every].append(inc)
        alt_summary_content[every].append(0)
        alt_summary_content[every].append('Card Name')
        alt_summary_content[every].append(value)
        alt_summary_content[every].append(inc)
        alt_summary_content[every].append(1)
        alt_summary_content[every].append('Part Number')
        alt_summary_content[every].append(value)
        alt_summary_content[every].append(inc)
        alt_summary_content[every].append(2)
        alt_summary_content[every].append('Product Name')
        alt_summary_content[every].append(value)
        alt_summary_content[every].append(inc)
        alt_summary_content[every].append(3)
        alt_summary_content[every].append('Date')
        alt_summary_content[every].append(value)
        alt_summary_content[every].append(inc)
        alt_summary_content[every].append(4)
        alt_summary_content[every].append('Version')
        alt_summary_content[every].append(value)
        alt_summary_content[every].append(inc)
        alt_summary_content[every].append(5)
        alt_summary_content[every].append('OS')
        alt_summary_content[every].append(value)
        alt_summary_content[every].append(inc)
        alt_summary_content[every].append(6)
        alt_summary_content[every].append('File name')
        alt_summary_content[every].append(value)
        alt_summary_content[every].append(inc)
        alt_summary_content[every].append(7)
        alt_summary_content[every].append('Download_URL')
        alt_summary_content[every].append(value)
        alt_summary_content[every].append(inc)
        alt_summary_content[every].append(8)
        alt_summary_content[every].append('Description')
        alt_summary_content[every].append(value)
        alt_summary_content[every].append(inc)
        alt_summary_content[every].append(9)
        alt_summary_content[every].append('Severity')
        alt_summary_content[every].append(value)
        
        for i in range (0,len(part_list)):
            flag=True
            value=False
            status=False
            
            for j in range (0,len(Product_name)):
                    
                if group_name[every]=='Mellanox OFED' and OFED_support[i+1]=='NO':
                    status=True
                    break
                elif group_name[every]=='WinOF' and WinOF_support[i+1]=='NO' :
                    status=True
                    break
                elif group_name[every]=='WinOF2' and WinOF2_support[i+1]=='NO':
                    status=True
                    break
                elif 'Mellanox MFT' in group_name[every] and MFT_support[i+1]=='NO':
                    status=True
                    break            
                elif 'ESXi' in group_name[every] and VM_support[i+1]=='NO':
                    status=True
                    break
                elif 'Windows firmware' in group_name[every] and Windows_fw_support[i+1]=='NO':
                    status=True
                    break
                elif 'Linux RoCE' in group_name[every] and Linux_RoCE_support[i+1]=='NO':
                    status=True
                    break
                elif 'Firmware binary' in group_name[every] and FW_binary_support[i+1]=='NO':
                    status=True
                    break
                elif search_elem[every] in Product_name[j] and part_list[i]==Part[j]:
                    inc+=1 
                    alt_summary_content[every].append(inc)
                    alt_summary_content[every].append(0)
                    
                    if flag==True:
                        alt_summary_content[every].append(card_name[i+1])
                        alt_summary_content[every].append('Column0')
                        alt_summary_content[every].append(inc)
                        alt_summary_content[every].append(1)
                        alt_summary_content[every].append(part_list[i])
                        flag=False
                    else:
                        alt_summary_content[every].append(' ')
                        alt_summary_content[every].append('Column0')
                        alt_summary_content[every].append(inc)
                        alt_summary_content[every].append(1)
                        alt_summary_content[every].append(' ')
                        
                    alt_summary_content[every].append('Column1')
                    alt_summary_content[every].append(inc)
                    alt_summary_content[every].append(2)
                    
                    if group_name[every]=='Mellanox OFED':
                        P = Product_name[j].split('for')
                        alt_summary_content[every].append(P[1])
                    else:
                        alt_summary_content[every].append(Product_name[j])
                        
                    if part_list[i].strip() not in Product_name[j] and group_name[every]=='Firmware binary posting':
                        alt_summary_content[every].append('Blue')
                    else:
                        alt_summary_content[every].append(value)
                        
                    alt_summary_content[every].append(inc)
                    alt_summary_content[every].append(3)
                    alt_summary_content[every].append(Date[j])
                    alt_summary_content[every].append(value)
                    alt_summary_content[every].append(inc)
                    alt_summary_content[every].append(4)
                    alt_summary_content[every].append(Version[j])
                    alt_summary_content[every].append(value)
                    alt_summary_content[every].append(inc)
                    alt_summary_content[every].append(5)
                    alt_summary_content[every].append(Os[j])
                    alt_summary_content[every].append(value)
                    alt_summary_content[every].append(inc)
                    alt_summary_content[every].append(6)
                    alt_summary_content[every].append(File_name[j])
                    
                    if File_name[j]=='Not Found':
                        alt_summary_content[every].append('Red_Bold')
                    else:
                        alt_summary_content[every].append(value) 
                        
                    alt_summary_content[every].append(inc)
                    alt_summary_content[every].append(7)
                    alt_summary_content[every].append(Download_page[j])
                    alt_summary_content[every].append(value)
                    alt_summary_content[every].append(inc)
                    alt_summary_content[every].append(8)
                    alt_summary_content[every].append(Description[j])
                    alt_summary_content[every].append(value)
                    alt_summary_content[every].append(inc)
                    alt_summary_content[every].append(9)
                    alt_summary_content[every].append(Severity[j])
                    alt_summary_content[every].append(value)
        
            if flag==True:
                inc+=1
                value = False
                alt_summary_content[every].append(inc)
                alt_summary_content[every].append(0)
                alt_summary_content[every].append(card_name[i+1])
                alt_summary_content[every].append('Column0')
                alt_summary_content[every].append(inc)
                alt_summary_content[every].append(1)
                alt_summary_content[every].append(part_list[i])
                alt_summary_content[every].append('Column1')
                alt_summary_content[every].append(inc)
                alt_summary_content[every].append(2)
                
                if status==True:
                    alt_summary_content[every].append('Not Supported')
                    alt_summary_content[every].append('Green_Bold')
                    status=False
                elif group_name[every]=='Firmware binary posting':
                    alt_summary_content[every].append('No firmware posting for '+part_list[i]+' found')
                    alt_summary_content[every].append('Red_Bold')
                else:
                    alt_summary_content[every].append('No Products Found')
                    alt_summary_content[every].append('Red_Bold')
    
                alt_summary_content[every].append(inc)
                alt_summary_content[every].append(3)
                alt_summary_content[every].append(' ')
                alt_summary_content[every].append(value)
                alt_summary_content[every].append(inc)
                alt_summary_content[every].append(4)
                alt_summary_content[every].append(' ')
                alt_summary_content[every].append(value)
                alt_summary_content[every].append(inc)
                alt_summary_content[every].append(5)
                alt_summary_content[every].append(' ')
                alt_summary_content[every].append(value)
                alt_summary_content[every].append(inc)
                alt_summary_content[every].append(6)
                alt_summary_content[every].append(' ')
                alt_summary_content[every].append(value)
                alt_summary_content[every].append(inc)
                alt_summary_content[every].append(7)
                alt_summary_content[every].append(' ')
                alt_summary_content[every].append(value)
                alt_summary_content[every].append(inc)
                alt_summary_content[every].append(8)
                alt_summary_content[every].append(' ')
                alt_summary_content[every].append(value)
                alt_summary_content[every].append(inc)
                alt_summary_content[every].append(9)
                alt_summary_content[every].append(' ')
                alt_summary_content[every].append(value)


def insert_input_sheet():
    input_sheet=[]
    global summary_content
    global alt_summary_content
    
    for i in range(0,len(Part2)):
        input_sheet.append(i)
        input_sheet.append(0)
        input_sheet.append(Part2[i])
        
        if i==0:
            input_sheet.append('Bold')
        else:
            input_sheet.append(False)
            
        input_sheet.append(i)
        input_sheet.append(1)
        input_sheet.append(card_name[i])
        
        if i==0:
            input_sheet.append('Bold')
        else:
            input_sheet.append(False)
            
        input_sheet.append(i)
        input_sheet.append(2)
        input_sheet.append(Chip[i])
        
        if i==0:
            input_sheet.append('Bold')
        else:
            input_sheet.append(False)
            
        input_sheet.append(i)
        input_sheet.append(3)
        input_sheet.append(Type[i])
        
        if i==0:
            input_sheet.append('Bold')
        else:
            input_sheet.append(False)
            
        input_sheet.append(i)
        input_sheet.append(4)
        input_sheet.append(WinOF_support[i])
        
        if i==0:
            input_sheet.append('Bold')
        else:
            input_sheet.append(False)
            
        input_sheet.append(i)
        input_sheet.append(5)
        input_sheet.append(WinOF2_support[i])
        
        if i==0:
            input_sheet.append('Bold')
        else:
            input_sheet.append(False)
            
        input_sheet.append(i)
        input_sheet.append(6)
        input_sheet.append(OFED_support[i])
        
        if i==0:
            input_sheet.append('Bold')
        else:
            input_sheet.append(False)
            
        input_sheet.append(i)
        input_sheet.append(7)
        input_sheet.append(VM_support[i])
        
        if i==0:
            input_sheet.append('Bold')
        else:
            input_sheet.append(False)
            
        input_sheet.append(i)
        input_sheet.append(8)
        input_sheet.append(MFT_support[i])
        
        if i==0:
            input_sheet.append('Bold')
        else:
            input_sheet.append(False)
            
        input_sheet.append(i)
        input_sheet.append(9)
        input_sheet.append(Windows_fw_support[i])
        
        if i==0:
            input_sheet.append('Bold')
        else:
            input_sheet.append(False)
            
        input_sheet.append(i)
        input_sheet.append(10)
        input_sheet.append(Linux_RoCE_support[i])
        
        if i==0:
            input_sheet.append('Bold')
        else:
            input_sheet.append(False)
            
        input_sheet.append(i)
        input_sheet.append(11)
        input_sheet.append(FW_binary_support[i])
        
        if i==0:
            input_sheet.append('Bold')
        else:
            input_sheet.append(False)
    
    summary_content.insert(0,input_sheet)
    alt_summary_content.insert(0,input_sheet)


output_path                                 =   get_output_path()
input_file , summary_output_file , alt_summary_output_file2    =   create_output_files(output_path)
search_elem, group_names, short_group_names =   get_keywords()

print(search_elem)

Part_numbers        =   extract(input_file,0)               # All the part numbers from unique report
Product_name        =   extract(input_file,1)               
Date                =   extract(input_file,2)
Version             =   extract(input_file,3)
Os                  =   extract(input_file,4)
File_name           =   extract(input_file,5)
Download_page       =   extract(input_file,6)
Description         =   extract(input_file,7)
Severity            =   extract(input_file,8)

Input_part_numbers  =   extract(sys.argv[1],0)              # All the part numbers from Audit Input file
Card_name           =   extract(sys.argv[1],1)
Chipset             =   extract(sys.argv[1],2)
Type                =   extract(sys.argv[1],3)
WinOF_support       =   extract(sys.argv[1],4)
WinOF2_support      =   extract(sys.argv[1],5)
OFED4_support       =   extract(sys.argv[1],6)
OFED5_support       =   extract(sys.argv[1],7)
VM_support          =   extract(sys.argv[1],8)
MFT_support         =   extract(sys.argv[1],9)
Windows_fw_support  =   extract(sys.argv[1],10)
Linux_RoCE_support  =   extract(sys.argv[1],11)
FW_binary_support   =   extract(sys.argv[1],12)
#FWPKG_support       =   extract(sys.argv[1],13)

summary_content     =   []
alt_summary_content =   []
part_number_list    =   []      
group_name          =   []
search_elem         =   []
short_group_name    =   []


if __name__=='__main__':
    logger = get_logger()
    logger.debug("Fetching part number and keyword froms search file")
    logger.debug("Summarizing into excel sheet with each part number as separate sheets")
    summary()
    logger.debug("Summarizing into excel sheet with each group as separate sheets")
    summary_alt()
    logger.debug("Copying input file to the output file for reference")
    insert_input_sheet()
    
    obj1=ExcelUtility(0)
    obj2=ExcelUtility(1)
    sheet_list1=[]
    sheet_list2=[]
    sheet_list1.append('input')
    sheet_list2.append('input')
    
    for i in range (0,len(part_list)):
        sheet_list1.append(part_list[i])
    
    for i in range (0,len(short_group_name)):
        sheet_list2.append(short_group_name[i])

    obj1.write_to_excel_file(output_file1,sheet_list1,summary_content)
    obj2.write_to_excel_file(output_file2,sheet_list2,alt_summary_content)
    
    logger.info("Summary file = "+str(output_file1))
    logger.info("Alternative summary file = "+str(output_file2))
    
    print("\n Output files are stored in the folder ' "+output_path+'\\Audit_'+sys.argv[2]+" '")